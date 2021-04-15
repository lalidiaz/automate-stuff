import openpyxl

# Read our spreadsheet file.
inv_file = openpyxl.load_workbook("inventory.xlsx")

# Get the specific sheet I will work on.
product_list = inv_file["Sheet1"]

# Calculate how many products I have per supplier and list the name of
# the suppliers with that respective number of products.
products_per_supplier = {}
total_value_per_supplier = {}
products_under_10_inv = {}

print(product_list.max_row)
# Read how many products are in the list to iterate.
# start from 2 and end in the last row (including)
for product_row in range(2, product_list.max_row + 1):
   supplier_name = product_list.cell(product_row, 4).value
   inventory = product_list.cell(product_row, 2).value
   price = product_list.cell(product_row, 3).value
   product_number = product_list.cell(product_row, 1).value
   inventory_price = product_list.cell(product_row, 5)

   # calculation for number of products per supplier:
   if supplier_name in products_per_supplier:
      current_num_products = products_per_supplier.get(supplier_name)
      products_per_supplier[supplier_name] = current_num_products + 1
   else:
      print("Adding a new supplier")
      products_per_supplier[supplier_name] = 1

   # calculate the total inventory value per supplier:
   if supplier_name in total_value_per_supplier:
      current_total_value = total_value_per_supplier.get(supplier_name)
      total_value_per_supplier[supplier_name] = current_total_value + inventory * price
   else:
      total_value_per_supplier[supplier_name] = inventory * price

   # logic products with inventory less than 10:
   if inventory < 10:
      products_under_10_inv[int(product_number)] = int(inventory)

   # add value for total inventory price:
   inventory_price.value = inventory * price


print(products_per_supplier)
print(total_value_per_supplier)
print(products_under_10_inv)
# save the file, it is going to create a new file:
inv_file.save("inventory_with_total_value.xlsx")










